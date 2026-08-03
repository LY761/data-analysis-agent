"""
数据导出服务 — 查询结果导出为 Excel 文件
"""
import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _available = True
except ImportError:
    _available = False


def export_to_excel(data: list, columns: list, title: str = "查询结果") -> bytes:
    """
    将查询结果导出为 Excel 文件。
    返回: Excel 文件的二进制内容
    """
    if not _available:
        raise RuntimeError("openpyxl 未安装，请执行: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据"

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=14, bold=True, color="1F2937")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # 副标题（导出时间）
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    time_cell = ws.cell(row=2, column=1, value=f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    time_cell.font = Font(size=10, color="9CA3AF")
    time_cell.alignment = Alignment(horizontal="center")

    # 表头
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    ws.row_dimensions[4].height = 25

    # 数据行
    for row_idx, row in enumerate(data, 5):
        for col_idx, col_name in enumerate(columns, 1):
            val = row.get(col_name, "")
            if isinstance(val, float):
                val = round(val, 2)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    # 自动列宽
    for col_idx, col_name in enumerate(columns, 1):
        max_width = max(len(str(col_name)), 12)
        for row in data[:50]:  # 只看前50行估算
            val_len = len(str(row.get(col_name, "")))
            max_width = max(max_width, min(val_len, 40))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_width + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info(f"[Export] Excel生成: {len(data)}行 x {len(columns)}列")
    return output.getvalue()


def is_available() -> bool:
    return _available
